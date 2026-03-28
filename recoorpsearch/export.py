"""
export.py — Export enriched PubMed records to Excel (.xlsx) or CSV.
"""

import os
from datetime import date
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

RESULTS_COLUMNS = [
    "pmid", "doi", "pubmed_url", "title", "journal", "pub_year", "pub_month",
    "author_name", "author_affiliation", "corporate_flag", "company_detected",
    "link_type", "detection_method", "coi_statement", "coi_corporate_flag",
    "coi_companies", "coi_link_type", "funding_sources", "private_funding_flag",
    "private_funders", "abstract",
]

RESULTS_HEADERS = {
    "pmid": "PMID", "doi": "DOI", "pubmed_url": "Link PubMed",
    "title": "T\u00edtulo", "journal": "Peri\u00f3dico", "pub_year": "Ano",
    "pub_month": "M\u00eas", "author_name": "Autor", "author_affiliation": "Afilia\u00e7\u00e3o",
    "corporate_flag": "V\u00ednculo Corporativo?", "company_detected": "Empresa Detectada",
    "link_type": "Tipo de V\u00ednculo", "detection_method": "M\u00e9todo de Detec\u00e7\u00e3o",
    "coi_statement": "Declara\u00e7\u00e3o de Conflito de Interesse",
    "coi_corporate_flag": "CoI Corporativo?", "coi_companies": "Empresas no CoI",
    "coi_link_type": "Tipo CoI", "funding_sources": "Fontes de Financiamento",
    "private_funding_flag": "Financiamento Privado?", "private_funders": "Financiadores Privados",
    "abstract": "Resumo",
}


def export_to_excel(
    records: list[dict],
    query: str,
    output_dir: str = ".",
    filename: Optional[str] = None,
    known_companies: Optional[list[str]] = None,
) -> str:
    df = _build_dataframe(records)
    today = date.today().isoformat()
    if not filename:
        slug = _slugify(query)[:40]
        filename = f"recoorpsearch_{slug}_{today}.xlsx"

    path = os.path.join(output_dir, filename)
    summary_df = _build_summary(df)
    params_df = _build_params(query, records, known_companies, today)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df[RESULTS_COLUMNS].rename(columns=RESULTS_HEADERS).to_excel(
            writer, sheet_name="Resultados", index=False
        )
        summary_df.to_excel(writer, sheet_name="Resumo por Empresa", index=False)
        params_df.to_excel(writer, sheet_name="Par\u00e2metros", index=False)

    _format_workbook(path)
    return os.path.abspath(path)


def export_to_csv(
    records: list[dict],
    query: str,
    output_dir: str = ".",
    filename: Optional[str] = None,
) -> str:
    df = _build_dataframe(records)
    today = date.today().isoformat()
    if not filename:
        slug = _slugify(query)[:40]
        filename = f"recoorpsearch_{slug}_{today}.csv"
    path = os.path.join(output_dir, filename)
    df[RESULTS_COLUMNS].rename(columns=RESULTS_HEADERS).to_csv(path, index=False, encoding="utf-8-sig")
    return os.path.abspath(path)


def _build_dataframe(records: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(records)
    for col in RESULTS_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df


def _build_summary(df: pd.DataFrame) -> pd.DataFrame:
    if "company_detected" not in df.columns or df.empty:
        return pd.DataFrame(columns=["Empresa", "N\u00ba Autores", "N\u00ba Artigos", "Peri\u00f3dicos", "Anos"])

    corp = df[df["corporate_flag"] == True].copy()
    if corp.empty:
        return pd.DataFrame(columns=["Empresa", "N\u00ba Autores", "N\u00ba Artigos", "Peri\u00f3dicos", "Anos"])

    corp["company_detected"] = corp["company_detected"].replace("", pd.NA).fillna("(n\u00e3o identificada)")

    grouped = corp.groupby("company_detected").agg(
        n_authors=("author_name", "nunique"),
        n_articles=("pmid", "nunique"),
        journals=("journal", lambda x: ", ".join(sorted(x.dropna().unique())[:5])),
        years=("pub_year", lambda x: _year_range(x)),
    ).reset_index()

    grouped.columns = ["Empresa", "N\u00ba Autores", "N\u00ba Artigos", "Peri\u00f3dicos", "Anos"]
    return grouped.sort_values("N\u00ba Artigos", ascending=False)


def _build_params(
    query: str,
    records: list[dict],
    known_companies: Optional[list[str]],
    today: str,
) -> pd.DataFrame:
    pmids = list({r.get("pmid", "") for r in records if r.get("pmid")})
    return pd.DataFrame(
        [
            {"Par\u00e2metro": "Query PubMed", "Valor": query},
            {"Par\u00e2metro": "Data da Busca", "Valor": today},
            {"Par\u00e2metro": "Total de PMIDs", "Valor": len(pmids)},
            {"Par\u00e2metro": "Total de Linhas (autor \u00d7 artigo)", "Valor": len(records)},
            {"Par\u00e2metro": "Empresas-alvo", "Valor": ", ".join(known_companies) if known_companies else "\u2014"},
            {"Par\u00e2metro": "Artigos com v\u00ednculo corporativo", "Valor": sum(1 for r in records if r.get("corporate_flag"))},
        ]
    )


_HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
_CORP_FILL   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_BODY_FONT   = Font(name="Arial", size=9)
_THIN        = Side(style="thin", color="BFBFBF")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _format_workbook(path: str) -> None:
    wb = load_workbook(path)
    for sheet in wb.worksheets:
        _format_sheet(sheet)

    ws = wb["Resultados"]
    corp_col_idx = None
    for cell in ws[1]:
        if cell.value == RESULTS_HEADERS["corporate_flag"]:
            corp_col_idx = cell.column
            break

    if corp_col_idx:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            flag_cell = row[corp_col_idx - 1]
            if str(flag_cell.value).upper() in ("TRUE", "1", "YES", "SIM"):
                for cell in row:
                    cell.fill = _CORP_FILL

    wb.save(path)


def _format_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    ws.row_dimensions[1].height = 36

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = _BORDER

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"


def _slugify(text: str) -> str:
    import re
    text = re.sub(r'[^\w\s-]', '', text.lower())
    return re.sub(r'[\s_-]+', '-', text).strip('-')


def _year_range(series: pd.Series) -> str:
    years = series.dropna().astype(str)
    years = [y for y in years if y.isdigit()]
    if not years:
        return "\u2014"
    return f"{min(years)}\u2013{max(years)}"
